from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from maskflow.core.engine import MaskingEngine
from maskflow.core.types import AnalysisResult


@dataclass(slots=True)
class _XlsxStats:
    matches_found: int = 0
    matches_applied: int = 0
    matches_skipped: int = 0
    detector_counts: Counter[str] = field(default_factory=Counter)
    detector_timings_ms: Counter[str] = field(default_factory=Counter)


class XlsxProcessor:
    def __init__(self, engine: MaskingEngine) -> None:
        self.engine = engine

    def analyze(self, source: Path | str) -> AnalysisResult:
        workbook = load_workbook(str(source))

        total_matches_found = 0
        total_matches_applied = 0
        total_matches_skipped = 0
        detector_counts: Counter[str] = Counter()
        detector_timings_ms: Counter[str] = Counter()

        for sheet in workbook.worksheets:
            if not isinstance(sheet, Worksheet):
                continue

            for row in sheet.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue

                    if cell.value.startswith("="):
                        continue

                    analysis = self.engine.analyze_text(cell.value)

                    total_matches_found += analysis.matches_found
                    total_matches_applied += analysis.matches_applied
                    total_matches_skipped += analysis.matches_skipped
                    detector_counts.update(analysis.detector_counts)
                    detector_timings_ms.update(analysis.detector_timings_ms)

        return AnalysisResult(
            matches_found=total_matches_found,
            matches_applied=total_matches_applied,
            matches_skipped=total_matches_skipped,
            detector_counts=dict(detector_counts),
            detector_timings_ms=dict(detector_timings_ms),
        )

    def process(self, source: Path | str, destination: Path | str) -> None:
        workbook = load_workbook(str(source))

        for sheet in workbook.worksheets:
            if not isinstance(sheet, Worksheet):
                continue

            for row in sheet.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue

                    if cell.value.startswith("="):
                        continue

                    cell.value = self.engine.process_text(cell.value)

        workbook.save(str(destination))

    def process_with_stats(
        self,
        source: Path | str,
        destination: Path | str,
    ) -> AnalysisResult:
        """Single-pass: маскирует и возвращает статистику за один проход.

        Устраняет двойное чтение файла по сравнению с analyze() + process().
        """
        workbook = load_workbook(str(source))
        stats = _XlsxStats()

        for sheet in workbook.worksheets:
            if not isinstance(sheet, Worksheet):
                continue

            for row in sheet.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue

                    if cell.value.startswith("="):
                        continue

                    masked, analysis = self.engine.process_with_stats(cell.value)
                    cell.value = masked

                    stats.matches_found += analysis.matches_found
                    stats.matches_applied += analysis.matches_applied
                    stats.matches_skipped += analysis.matches_skipped
                    stats.detector_counts.update(analysis.detector_counts)
                    stats.detector_timings_ms.update(analysis.detector_timings_ms)

        workbook.save(str(destination))

        return AnalysisResult(
            matches_found=stats.matches_found,
            matches_applied=stats.matches_applied,
            matches_skipped=stats.matches_skipped,
            detector_counts=dict(stats.detector_counts),
            detector_timings_ms=dict(stats.detector_timings_ms),
        )
